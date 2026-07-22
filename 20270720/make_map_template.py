# -*- coding: utf-8 -*-
"""生成「供需协同地图 - 悬浮框基础数据模板」Excel，交付给后端开发。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------- 样式 ----------
TITLE = Font(name="微软雅黑", size=14, bold=True, color="1F2430")
H1 = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
NORMAL = Font(name="微软雅黑", size=10, color="1F2430")
NOTE = Font(name="微软雅黑", size=9, color="6B7280", italic=True)
RED = Font(name="微软雅黑", size=10, bold=True, color="B91C1C")

HEAD_FILL = PatternFill("solid", fgColor="2563EB")
HEAD_FILL_RED = PatternFill("solid", fgColor="D6001C")
HEAD_FILL_GRAY = PatternFill("solid", fgColor="374151")
STATIC_FILL = PatternFill("solid", fgColor="EEF2FF")
DYN_FILL = PatternFill("solid", fgColor="FEF3C7")

thin = Side(style="thin", color="D7DBE2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, ncols, fill=HEAD_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H1
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER


def fill_rows(ws, start_row, data, ncols, body_fill=None):
    for i, rowvals in enumerate(data):
        r = start_row + i
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c, value=rowvals[c - 1])
            cell.font = NORMAL
            cell.alignment = WRAP
            cell.border = BORDER
            if body_fill:
                cell.fill = body_fill


def set_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ============================================================
# Sheet 0 · 说明
# ============================================================
ws = wb.active
ws.title = "说明"
set_widths(ws, [22, 26, 42, 26])
ws["A1"] = "供需协同地图 · 悬浮框基础数据模板"
ws["A1"].font = TITLE
ws.merge_cells("A1:D1")
ws["A3"] = "用途：本模板用于向后端开发交付「供需协同地图」悬浮框所需的基础数据与字段定义。"
ws["A3"].font = NORMAL
ws.merge_cells("A3:D3")
ws["A4"] = "颜色约定：蓝色=静态主数据(人工维护)；黄色=动态数据(随「上传数据」上传，系统展示最新版本)。"
ws["A4"].font = NOTE
ws.merge_cells("A4:D4")

ws["A6"] = "地图上共有两类悬浮框，数据来源如下："
ws["A6"].font = Font(name="微软雅黑", size=11, bold=True)
ws.merge_cells("A6:D6")

head = ["悬浮框", "触发位置", "展示字段", "数据来源"]
for c, v in enumerate(head, 1):
    ws.cell(row=7, column=c, value=v)
style_header(ws, 7, 4, HEAD_FILL_GRAY)

rows = [
    ["A. 连线悬浮框", "悬浮在协同连线上",
     "产品 / 信息流 / 实物流（静态）；风险零件（动态·随上传·最新版本）",
     "① 协同关系主表 + ② 风险零件上传模板"],
    ["B. 节点悬浮框", "悬浮在 Entity 坐标点上",
     "未来 6 个月最新版预测；经纬度",
     "已有模块：预测=「上传数据」；坐标=「组织地址坐标」，本模板不重复提供"],
]
fill_rows(ws, 8, rows, 4)
ws.row_dimensions[8].height = 46
ws.row_dimensions[9].height = 46

ws["A12"] = "需要你提供的数据 = ① 协同关系主表（静态） + ② 风险零件上传模板（动态）。"
ws["A12"].font = RED
ws.merge_cells("A12:D12")
ws["A13"] = "坐标、预测数据来自已有模块，后端直接从系统取，无需在本文件提供。"
ws["A13"].font = NOTE
ws.merge_cells("A13:D13")

# ============================================================
# Sheet 1 · 协同关系主表（静态）
# ============================================================
ws1 = wb.create_sheet("①协同关系主表(静态)")
set_widths(ws1, [12, 14, 14, 22, 26, 26, 10, 20])
ws1["A1"] = "① 供需协同关系主表（静态主数据，人工维护，变动不频繁）"
ws1["A1"].font = TITLE
ws1.merge_cells("A1:H1")

cols1 = ["关系ID*", "供方Entity*", "需方Entity*", "产品*", "信息流*", "实物流*", "状态", "备注"]
for c, v in enumerate(cols1, 1):
    ws1.cell(row=3, column=c, value=v)
style_header(ws1, 3, len(cols1), HEAD_FILL)

data1 = [
    ["R001", "CCS-WH", "XCEC", "重型发动机总成", "需求预测 / 排产计划", "整机 & 关键零部件", "启用", "示例数据，可删除"],
    ["R002", "DCEC", "ACPL", "轻型发动机", "需求预测 / 库存共享", "整机", "启用", ""],
    ["", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", ""],
]
fill_rows(ws1, 4, data1, len(cols1), STATIC_FILL)
ws1["A9"] = "注：关系ID 为连线唯一标识；供方/需方Entity 请填「组织地址坐标」表中的简称，保证能匹配到坐标。带*为必填。"
ws1["A9"].font = NOTE
ws1.merge_cells("A9:H9")

# ============================================================
# Sheet 2 · 风险零件上传模板（动态）
# ============================================================
ws2 = wb.create_sheet("②风险零件上传模板(动态)")
set_widths(ws2, [20, 14, 14, 18, 40, 10, 18])
ws2["A1"] = "② 风险零件上传模板（动态数据，随「上传数据」上传；地图悬浮展示最新版本）"
ws2["A1"].font = TITLE
ws2.merge_cells("A1:G1")

cols2 = ["上传版本(Meeting Date)*", "供方Entity*", "需方Entity*", "零件名称*", "风险描述*", "风险等级", "关系ID"]
for c, v in enumerate(cols2, 1):
    ws2.cell(row=3, column=c, value=v)
style_header(ws2, 3, len(cols2), HEAD_FILL_RED)

data2 = [
    ["2026-05-15", "CCS-WH", "XCEC", "高压共轨泵", "供应周期偏长，需提前8周锁定", "高", "R001"],
    ["2026-05-15", "CCS-WH", "XCEC", "ECU芯片", "存在缺货风险", "高", "R001"],
    ["2026-04-15", "CCS-WH", "XCEC", "高压共轨泵", "（上一版本，仅历史留存，地图不展示）", "中", "R001"],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
]
fill_rows(ws2, 4, data2, len(cols2), DYN_FILL)
ws2["A10"] = "注：同一供需关系可有多个风险零件（一对多）；同一关系在不同上传版本(Meeting Date)会有多条，地图悬浮只取【最新 Meeting Date】的那一版。带*为必填。"
ws2["A10"].font = NOTE
ws2.merge_cells("A10:G10")

# ============================================================
# Sheet 3 · 字段说明
# ============================================================
ws3 = wb.create_sheet("字段说明")
set_widths(ws3, [22, 22, 10, 10, 30, 26])
ws3["A1"] = "字段说明 / 数据字典"
ws3["A1"].font = TITLE
ws3.merge_cells("A1:F1")

cols3 = ["所属表", "字段", "类型", "必填", "示例", "说明 / 来源"]
for c, v in enumerate(cols3, 1):
    ws3.cell(row=3, column=c, value=v)
style_header(ws3, 3, len(cols3), HEAD_FILL_GRAY)

data3 = [
    ["①协同关系主表", "关系ID", "文本", "是", "R001", "主键，连线唯一标识"],
    ["①协同关系主表", "供方Entity", "文本", "是", "CCS-WH", "关联组织地址坐标表的简称"],
    ["①协同关系主表", "需方Entity", "文本", "是", "XCEC", "关联组织地址坐标表的简称"],
    ["①协同关系主表", "产品", "文本", "是", "重型发动机总成", "连线悬浮展示"],
    ["①协同关系主表", "信息流", "文本", "是", "需求预测 / 排产计划", "连线悬浮展示"],
    ["①协同关系主表", "实物流", "文本", "是", "整机 & 关键零部件", "连线悬浮展示"],
    ["①协同关系主表", "状态", "枚举", "否", "启用/停用", "停用则地图不显示该连线"],
    ["②风险零件", "上传版本(Meeting Date)", "日期", "是", "2026-05-15", "上传时的版本；地图取最新版本"],
    ["②风险零件", "供方Entity", "文本", "是", "CCS-WH", "与协同关系匹配"],
    ["②风险零件", "需方Entity", "文本", "是", "XCEC", "与协同关系匹配"],
    ["②风险零件", "零件名称", "文本", "是", "高压共轨泵", "连线悬浮·风险零件"],
    ["②风险零件", "风险描述", "文本", "是", "供应周期偏长，需提前8周锁定", "连线悬浮·风险零件（可较长）"],
    ["②风险零件", "风险等级", "枚举", "否", "高/中/低", "可用于悬浮标红/排序"],
    ["②风险零件", "关系ID", "文本", "否", "R001", "外键，关联①主表；便于精确匹配"],
]
fill_rows(ws3, 4, data3, len(cols3))
for r in range(4, 4 + len(data3)):
    ws3.row_dimensions[r].height = 22

for w in [ws, ws1, ws2, ws3]:
    w.sheet_view.showGridLines = False

out = r"f:\cursor\test01\20270720\供需协同地图_悬浮数据模板.xlsx"
wb.save(out)
print("saved:", out)
